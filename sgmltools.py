import openpyxl
import xlrd
import csv
from bs4 import BeautifulSoup
from tabulate import tabulate
import re
import logging

class SGMLtools:
    """Master class to work with Airbus SGML files"""

    def __init__(self, source_file):
        """Load the source data"""
        self.log = logging
        self.log.basicConfig(level=logging.INFO)
        self.log.debug('Opening {}'.format(source_file))
        with open(source_file, 'r') as source_data:
            data = source_data.read()
            self.manual = BeautifulSoup(data, 'html.parser')
            self.log.debug('Parsing succesful')
        self.log.info('AC model: {}'.format(self.manual.amm['model']))
        self.log.debug('Init OK')

    def findAllTasks(self, recurse_limit=None):
        """Searach the Manual for anything with a task tag

        Returns a list
        """
        return self.manual.find_all('task', limit=recurse_limit)

    def extractTaskInfo(self, task):
        """Get all info from a task"""
        taskcontent = dict()
        title = task.title.string
        self.log.info('TASK: {}'.format(title))
        refstring = self.__splitCode(task)
        self.log.debug(refstring)
        topics = task.find_all("topic")
        for x in range(0, len(topics)):
            subtasks = topics[x].find_all("subtask")
            for y in range(0, len(subtasks)):
                subtask = dict()
##                subtasks[y].effect['effrg']
                try:
                    effec = subtasks[y].effect['effrg']
                except TypeError:
                    self.log.warning('TypeError on {}'.format(title))
                    continue
##                self.log.debug("Effectivity: %s\nAccomplish A320 AMM %s Subtask %s: %s" % (effec,
##                                                                                           topics[x].title.string,
##                                                                                           self.__splitCode(subtasks[y]),
####                                                                                           subtask_string(subtasks[y]),
##                                                                                           subtasks[y].para.string))
                if subtasks[y].find('list1') is not None:
                    self.getTaskContent(subtasks[y].list1,1)
                if subtasks[y].find("cblst") != None:
                    cblst = True

    def findListObject(self, parent_obj):
        """Extract the lists"""
        recursion_level = 1
        listcaptureRE = re.compile('list(\d)', re.IGNORECASE)
        lists = parent_obj.find_all(listcaptureRE)
        for item in lists:
            matchobj = re.match(listcaptureRE, item.name)
            if matchobj is not None:
                if int(matchobj.group(1)) > recursion_level:
                    recursion_level = int(matchobj.group(1))
        return str(recursion_level)

    def getTaskContent(self, parent_obj, level):
        """Walk through the task and get the info"""
##        self.log.warning('LEVEL INPUT {}'.format(level))
        listcaptureRE = re.compile('list(\d)', re.IGNORECASE)
        list_items = parent_obj.find_all('l{}item'.format(level), recursive=False)
        for item in list_items:
##            self.log.debug('Para {}'.format(item.para.string))
            for nextitem in item:
                self.log.debug('Parent:{} This:{} Content:{}'.format(nextitem.parent.name, nextitem.name, nextitem.string))
                if nextitem.para is None:
                    self.log.debug("Childdata")
                    listitemcontent = list()
                    for space in range(int(level)):
                        listitemcontent.append(' ')
                    for child in nextitem.descendants:
                        self.log.debug('Parent:{} This:{} Content:{}'.format(child.parent.name, child.name, child.string))
                        if child.string is not None and child.parent.name not in ['pan', 'refint', 'con', 'std', 'stdname']:
                            string_ob = str(child.string).rstrip()
                            string_ob = string_ob.lstrip()
                            listitemcontent.append(string_ob)
                    while '' in listitemcontent:
                        listitemcontent.remove('')
                    self.log.info(' '.join(self.__fixPunctuation(listitemcontent)))
##                    print(listitemcontent)
            if item.find(listcaptureRE) is not None:
                nextlevel = re.match(listcaptureRE, item.find(listcaptureRE).name)
                self.getTaskContent(item.find(listcaptureRE), nextlevel.group(1))
        
    def __fixPunctuation(self, list_to_fix):
        """Private method to fix the punctuation"""
        punclist = ['.', ',']
        for punc in punclist:
            while punc in list_to_fix:
                try:
                    index_pos = list_to_fix.index(punc)
                except ValueError:
                    continue
                popped_item = list_to_fix.pop(index_pos)
                list_to_fix[index_pos - 1] = list_to_fix[index_pos - 1] + popped_item
        return list_to_fix
        
    def __splitCode(self, tag):
        """Private method to split the task code into into its major parts"""
        tagattrs = tag.attrs
        if 'varnbr' in tagattrs:
            var = tagattrs['varnbr']
        else:
            var = ''
        return '-'.join((tagattrs['chapnbr'],
                         tagattrs['sectnbr'],
                         tagattrs['subjnbr'],
                         tagattrs['func'],
                         tagattrs['seq'],
                         tagattrs['confltr']+var
                         ))



    def getSteps(self, task, zone):
        topics = task.find_all("topic")
        print(task.title.string)
        side = get_side(zone)
        if side is not None:
            print(side)
    ##    subtasks = task.find_all("subtask")
        for x in range(0, len(topics)):
            subtasks = topics[x].find_all("subtask")
            for y in range(0, len(subtasks)):
                print("Effectivity: %s\nAccomplish A320 AMM %s Subtask %s: %s" % (subtasks[y].effect['effrg'], topics[x].title.string, subtask_string(subtasks[y]), subtasks[y].para.string))
                if subtasks[y].find("cblst") != None:
                    cblst = subtasks[y].find("cblst")
                    cbdata = cblst.find_all("cbdata")
                    hdrs = ["FIN", "Name", "Panel", "Location", "Effectivity"]
                    listofbreakers = []
                    for z in range(0, len(cbdata)):
                        row = []
                        row.append(cbdata[z].find("cb").string)
                        row.append(cbdata[z].find("cbname").string)
                        row.append(cbdata[z].find("pan").string)
                        row.append(cbdata[z].find("cbloc").string)
                        effect = cbdata[z].find("effect")
                        row.append(effect.attrs['effrg'])
                        listofbreakers.append(row)
                    print(tabulate(listofbreakers, hdrs, tablefmt="grid"))
    ##        print(topics[x].title.string)

    

if __name__ == '__main__':
    ddata = SGMLtools('d:/Code/A320/AMM.SGM')
    boop = ddata.findAllTasks(recurse_limit=250)
    for things in boop:
        ddata.extractTaskInfo(things)



##
##ammfile = open("AMM.SGM",'r')
##data = ammfile.read()
##amm = BeautifulSoup(data, 'html.parser')
##print("Imported AMM SGML file")
##
##mpdfile = open("MPD.SGM", "r")
##mpddata = mpdfile.read()
##mpd = BeautifulSoup(mpddata, 'html.parser')
##print("Imported MPD SGML file")
##
##temfile = open("TEM.SGM", 'r')
##temdata = temfile.read()
##tem = BeautifulSoup(temdata, 'html.parser')
##print("Imported TEM SGML file")
##
##outputbook = openpyxl.Workbook()
##task2toolsheet = outputbook.active
##task2toolsheet.title = "Tooling Requirements"
##STDtoolsheet = outputbook.create_sheet("Standard Tools")
##skippedtasksheet = outputbook.create_sheet("Skipped Tasks")
##task2toolsheet['a1'].value = "MPD/JQ Task Number"
##task2toolsheet['b1'].value = "MPD Title"
##task2toolsheet['c1'].value = "AMM Task"
##task2toolsheet['d1'].value = "Tool Number"
##task2toolsheet['e1'].value = "Tool Description"
##
##STDtoolsheet['a1'].value = "MPD/JQ Task Number"
##STDtoolsheet['b1'].value = "MPD Title"
##STDtoolsheet['c1'].value = "AMM Task"
####STDtoolsheet['d1'].value = "Tool Number"
##STDtoolsheet['d1'].value = "Tool Description"
##
##
##skippedtasksheet['a1'].value = "Input sheet Line Number"
##skippedtasksheet['b1'].value = "JQ Task Number"
##skippedtasksheet['c1'].value = "Skip Reason"
##
##print("Spawned Output Object")
##
##print("Ready...")
##
##zones = amm.zoning
##
##def get_side(zone):
##    zonetag = zones.find("zone", string=zone)
##    zonedata = zonetag.find_next_sibling("zondata")
##    return zonedata.side.attrs['hand']
##
##def find_task_in_MPD(task_cd):
##    taskinfo = {}
##    for x in range(3, mpdsheet.nrows):
##        if mpdsheet.cell(x, 2).value == task_cd:
##            taskinfo = {'access': mpdsheet.cell(x, 4).value, 'zone': mpdsheet.cell(x, 6).value, 'description': mpdsheet.cell(x, 7).value, 'labourskill': mpdsheet.cell(x, 8).value, 'reference': (mpdsheet.cell(x, 17).value)[:14], 'labourqty':mpdsheet.cell(x, 18).value, 'labourmhours':mpdsheet.cell(x, 19).value}
##            return taskinfo
##            break
##
##def find_task_in_AMM(ref_cd):
##    a = amm.find("task", key=ref_cd+"00")
##    if a == None:
##        a = amm.find("task", key=ref_cd)
####    print(amm.find("task", key=ref_cd))
##    return a
##
##def mpd2key(reference):
##    newref = reference.replace("-", "")
##    return "EN"+newref
##
##def read_task(cellinfo):
####    cellinfo = tasklist['C'+str(row)].value
##    if cellinfo.find(" ") != -1:
##        taskID = cellinfo[:cellinfo.find(" ")]
##        zone = cellinfo[cellinfo.rfind(" ")+1:]
##        print("B")
##    else:
##        taskID = cellinfo
##        zone = None
##        print("A")
##    return taskID,zone
####    print(find_task_in_MPD(taskID))
##
##def task_from_MPD(desc):
##    taskname = desc.split("\n\n")
##    taskname = taskname[1]
##    taskname = taskname.replace("\n", "")
##    return taskname
##
##def AMTOSS_from_ref(ref):
##    firstsplit = ref[4:]
##    thirdsplit = ref[2:4]
##    secondsplit = ref[:2]
##    return secondsplit+"-"+thirdsplit+"-"+firstsplit
##
##def display_nicely(taskinfo):
##    print("Task: "+task_from_MPD(taskinfo['description']))
##    print("AMTOSS: "+AMTOSS_from_ref(taskinfo['reference']))
##    print("Zone: "+taskinfo['zone'])
##    print("Access: "+taskinfo['access'])
##    print("Trade: "+taskinfo['labourskill'])
##    print("Resources: "+taskinfo['labourqty'])
##    print("Man Hours: "+taskinfo['labourmhours'])
##    print("------------------------------")
##
##def subtask_string(subtasktag):
##    subtaskattrs = subtasktag.attrs
##    return "%s-%s-%s-%s-%s-%s" % (subtaskattrs['chapnbr'], subtaskattrs['sectnbr'], subtaskattrs['subjnbr'], subtaskattrs['func'], subtaskattrs['seq'], subtaskattrs['confltr'])
##
##def go():
##    for x in range(2, tasklist.max_row):
##        print(x)
##    ##    print(tasklist["C"+str(x)].value)
##        try:
##            mpdref = read_task(tasklist["C"+str(x)].value)[0]
####            print(mpdref)
##            mpddeets = find_task_in_MPD(mpdref)
####            print(mpddeets)
##            ammref = mpd2key(mpddeets['reference'])
##            print(ammref)
##            task = find_task_in_AMM(ammref)
##            get_steps(task, read_task(tasklist["C"+str(x)].value)[1])
####            display_nicely(find_task_in_MPD(tasklist["C"+str(x)].value))
##        except:
##            pass
####    a = find_task_in_MPD("532209-01-1")
####find_task_in_AMM(mpd2key(a['reference']))
##
####Append the tool data to the output sheet
##
##
##
####Break the JQ task into sections to search for the attributes in the SGML MPD
##def task2Attrs(task):
##    if re.match('^ZL', task):
##        return task[:6], task[7], task[8], task[10]
##    else:
##        return task[:2], task[2:4], task[4:6], task[8], task[10], task[7]
##
##
####Find all tooling information for every MPD task in a list.  Points at the JQ spreadsheet right now.
##def findToolsForTasks():
##    for linenumber in range(2, tasklist.max_row):
##        print('Spreadsheet Line: %s' %linenumber)
##        taskID = tasklist["B"+str(linenumber)].value
##        if re.match('(^[0-9]{6}-[ABI0-9][0-9]-[0-9])|^ZL', taskID):
##
##            print("Task: %s" % taskID)
##            taskattributes = task2Attrs(taskID)
####            print(taskattributes)
##            if re.match('^ZL', taskID):
##                smtask = mpd.find("smtask", zonalref=taskattributes[0], etid=taskattributes[1], seqnbr=taskattributes[2], solution=taskattributes[3])
##            else:
##                smtask = mpd.find("smtask", atach=taskattributes[0], atasect=taskattributes[1], atasubj=taskattributes[2], seqnbr=taskattributes[3], solution=taskattributes[4], etid=taskattributes[5])
##            if smtask is not None:
##                try:
##                    mpdtitle = smtask.find("title").string
##                    print(mpdtitle)
##                except AttributeError:
##                    print("Couldnt find the MPD task title.  Probably an MPD data struture problem")
##                ammlinks = smtask.find_all("refext", refman="AMM")
##                if len(ammlinks) == 0:
##                    print("MPD did not provide a valid AMM reference")
##                for ammref in ammlinks:
##    ##                print("AMM Key: %s" % mpd2key(ammref['refloc']))
##                    ammtask = amm.find("task", key=mpd2key(ammref['refloc'])+'00')
##                    if ammtask is not None:
##                        ammtitle = ammtask.title.string
##                        print(ammtitle)
##                        addedtools = list()
##                        addedstdtools = list()
##                        tools = ammtask.find_all("ted")
##                        if len(tools) >0 :
##                            for tedblock in tools:
##                                toolnumbers = tedblock.find_all("toolnbr")
##                                toolnames = tedblock.find_all("toolname")
##                                for x in range(0, len(toolnumbers)):
##                                    if toolnumbers[x].string not in addedtools:
##                                        print("Tool number: %s" % toolnumbers[x].string)
##                                        print("Tool name: %s" % toolnames[x].string)
##                                        addedtools.append(toolnumbers[x].string)
##                                        row = len(task2toolsheet['A'])+1
##                                        task2toolsheet['A'+str(row)].value = taskID
##                                        task2toolsheet['B'+str(row)].value = mpdtitle
##                                        task2toolsheet['C'+str(row)].value = ammtitle
##
##                                        task2toolsheet['D'+str(row)].value = toolnumbers[x].string
##                                        task2toolsheet['E'+str(row)].value = toolnames[x].string
##    ##                                    print("Supplier Code: %s" % getToolVendorData(toolnumbers[x].string))
##                        else:
##                            print("No Specific Tools")
####                            row = len(skippedtasksheet['A'])+1
####                            skippedtasksheet['A'+str(row)].value = linenumber
####                            skippedtasksheet['B'+str(row)].value = mpdtitle
####                            skippedtasksheet['C'+str(row)].value = "No Tools"
##
##                        stdtools = ammtask.find_all("std")
##                        if len(stdtools) > 0 :
##                            for stdblock in stdtools:
##                                stdtoolnames = stdblock.find_all("stdname")
##                                for x in range(0, len(stdtoolnames)):
##                                    print("Standard Tool: %s" % stdtoolnames[x].string)
##                                    row = len(STDtoolsheet['A'])+1
##                                    STDtoolsheet['A'+str(row)].value = taskID
##                                    STDtoolsheet['B'+str(row)].value = mpdtitle
##                                    STDtoolsheet['C'+str(row)].value = ammtitle
####                                    STDtoolsheet['D'+str(row)].value = toolnumbers[x].string
##                                    STDtoolsheet['D'+str(row)].value = stdtoolnames[x].string
##                        else:
##                            print("No Standard Tools")
##                    else:
##                        print("Did not find a matching AMM task")
##                        row = len(skippedtasksheet['A'])+1
##                        skippedtasksheet['A'+str(row)].value = linenumber
##                        skippedtasksheet['B'+str(row)].value = taskID
##                        skippedtasksheet['C'+str(row)].value = "Did not find a matching AMM task"
##            else:
##                print("Did not find a matching MPD task")
##                row = len(skippedtasksheet['A'])+1
##                skippedtasksheet['A'+str(row)].value = linenumber
##                skippedtasksheet['B'+str(row)].value = taskID
##                skippedtasksheet['C'+str(row)].value = "Did not find a matching MPD task"
##        else:
##            row = len(skippedtasksheet['A'])+1
##            skippedtasksheet['A'+str(row)].value = linenumber
##            skippedtasksheet['B'+str(row)].value = taskID
##            skippedtasksheet['C'+str(row)].value = "Is Not an MPD task"
##
##
##        print("---------------------------")
##
##
##def getToolVendorData(toolpartnum):
##    tools = tem.tem.toolst.find_all("tool")
##    for x in range(0, len(tools)):
####        print(tools[x].find("toolnbr"))
##        try:
##            toolinfo = tools[x].find("toolnbr").find("pnr", string=re.compile(toolpartnum))
##            if toolinfo is not None:
##                break
##        except AttributeError:
##            print(x)
####        print(toolinfo)
####    print(toolinfo)
##    firstparent = toolinfo.parent
##    secondparent = firstparent.parent
####    thirdparent = secondparent.parent
####    fourthparent = thirdparent.parent
####    print(secondparent.prettify())
##    suppliercode = secondparent.find("spl")
##    return suppliercode.string
####    vendordetails = tem.tem.mfmatr.vendlist.find("mfr", string=re.compile(suppliercode.string))
####    print(vendordetails)
######    print(secondparent.prettify())
##
##
##
##
##
##
####print(mpd2key(a['reference']))
##
####print(a)
##
##def get_all_zones():
##    counter = 1
##    with open('names.csv', 'wb') as csvfile:
##        fieldnames = ['zone_cd', 'zone_name', 'next_higher_zone_cd', 'side']
##        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
####        writer.writeheader()
##        for x in range(0, len(zones.find_all("majzone"))):
##            print("zone: %s"%zones.find_all("majzone")[x].zone.string)
##            print("zonedesc: %s"%zones.find_all("majzone")[x].zondata.zonedesc.string)
##            writer.writerow({'zone_cd':zones.find_all("majzone")[x].zone.string, 'zone_name':zones.find_all("majzone")[x].zondata.zonedesc.string.replace(u'\xb0', "o"), 'next_higher_zone_cd':'', 'side':zones.find_all("majzone")[x].zondata.side['hand']})
##            counter = counter +1
##            for y  in range(0, len(zones.find_all("majzone")[x].find_all("subzone"))):
##                    print("nexthigher: %s" % zones.find_all("majzone")[x].zone.string)
##                    print("zone: %s" % zones.find_all("majzone")[x].find_all("subzone")[y].zone.string)
##                    print("zonedesc: %s" % zones.find_all("majzone")[x].find_all("subzone")[y].zondata.zonedesc.string)
##                    writer.writerow({'zone_cd':zones.find_all("majzone")[x].find_all("subzone")[y].zone.string, 'zone_name':zones.find_all("majzone")[x].find_all("subzone")[y].zondata.zonedesc.string.replace(u'\xb0', "o"), 'next_higher_zone_cd':zones.find_all("majzone")[x].zone.string, 'side':zones.find_all("majzone")[x].find_all("subzone")[y].zondata.side['hand']})
##                    counter = counter +1
##                    for z in range(0, len(zones.find_all("majzone")[x].find_all("subzone")[y].find_all("speczone"))):
##                            print("nexthigher: %s" % zones.find_all("majzone")[x].zone.string)
##                            print("zone: %s" % zones.find_all("majzone")[x].find_all("subzone")[y].find_all("speczone")[z].zone.string)
##                            print("zonedesc: %s" % zones.find_all("majzone")[x].find_all("subzone")[y].find_all("speczone")[z].zondata.zonedesc.string)
##                            writer.writerow({'zone_cd':zones.find_all("majzone")[x].find_all("subzone")[y].find_all("speczone")[z].zone.string, 'zone_name':zones.find_all("majzone")[x].find_all("subzone")[y].find_all("speczone")[z].zondata.zonedesc.string.replace(u'\xb0', "o"), 'next_higher_zone_cd':zones.find_all("majzone")[x].zone.string, 'side':zones.find_all("majzone")[x].find_all("subzone")[y].find_all("speczone")[z].zondata.side['hand']})
##                            counter = counter +1
##
##def get_all_panels():
##    with open('panels.csv', 'wb') as panelfile:
##        fieldnames = ['panel']
##        writer = csv.DictWriter(panelfile, fieldnames=fieldnames)
##        for x in range(0, len(c)):
##            writer.writerow({'panel':c[x].string})
##
##def makeToolReport():
##    findToolsForTasks()
##    print("Saving Workbook")
##    outputbook.save('tooling.xlsx')
##    print("Finished.")
##go()
##
##makeToolReport()
